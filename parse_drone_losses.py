"""
parse_drone_losses.py

Parses Ukrainian combat drone loss reports (.docx) and writes
all records into a single .xlsx output file.

Usage: run parse_drone_losses.exe from the folder containing the .docx report files.

Build (one-time, technical):
    pip install python-docx openpyxl pyinstaller
    pyinstaller --onefile --console parse_drone_losses.py
"""

import re
import sys
import logging
from datetime import date, time
from pathlib import Path

from docx import Document
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Windows console UTF-8 (prevents garbled Ukrainian text in cmd/PowerShell)
# ---------------------------------------------------------------------------

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except AttributeError:
        pass

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------


def setup_logging() -> logging.Logger:
    logger = logging.getLogger("drone_parser")
    logger.setLevel(logging.DEBUG)
    handler = logging.StreamHandler(sys.stdout)
    handler.setFormatter(logging.Formatter("%(message)s"))
    logger.addHandler(handler)
    return logger


log = setup_logging()

# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------


def find_docx_files(folder: Path) -> list[Path]:
    """Return all .docx files in folder, skipping Word temp files (~$...)."""
    return sorted(f for f in folder.glob("*.docx") if not f.name.startswith("~$"))

# ---------------------------------------------------------------------------
# Docx text extraction
# ---------------------------------------------------------------------------


def extract_paragraphs(filepath: Path) -> list[str]:
    """Extract all non-empty paragraph texts from a .docx file."""
    doc = Document(str(filepath))
    return [p.text.strip() for p in doc.paragraphs if p.text.strip()]

# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------


def _find_inline_or_next(paragraphs: list[str], label: str) -> str | None:
    """
    Locate a field value by label. Handles two document patterns:
      - "Label: value"           -> returns 'value' (inline)
      - "Label:"                 -> returns the next non-empty paragraph
      - "Label (extra text):"    -> returns the next non-empty paragraph
    Matching is case-insensitive.
    """
    label_lower = label.lower()
    for i, para in enumerate(paragraphs):
        if label_lower not in para.lower():
            continue
        idx = para.lower().index(label_lower) + len(label)
        remainder = para[idx:].strip()
        # Strip leading separator characters to reveal any inline value
        inline = remainder.lstrip(":–—() \t").strip()
        # If the line ends with ':' it is a header-only line; value is on next paragraph
        if not inline or remainder.rstrip().endswith(":"):
            for j in range(i + 1, min(i + 3, len(paragraphs))):
                if paragraphs[j]:
                    return paragraphs[j]
            return None
        return inline
    return None


def parse_drone_model(paragraphs: list[str]) -> str | None:
    """
    Extract drone model from the ВТРАТА line.

    Handles:
      'ВТРАТА   БпЛА «DJI Mavic 4 Pro»'  ->  'DJI Mavic 4 Pro'
      'ВТРАТА   AUTEL EVO MAX 4Т'          ->  'AUTEL EVO MAX 4Т'
    """
    for para in paragraphs:
        if "ВТРАТА" not in para.upper():
            continue
        # Try Ukrainian guillemets «...» first
        m = re.search(r"«([^»]+)»", para)
        if m:
            return m.group(1).strip()
        # Fallback: strip ВТРАТА and БпЛА keywords, return the remainder
        model = re.sub(r"ВТРАТА", "", para, flags=re.IGNORECASE)
        model = re.sub(r"БпЛА", "", model, flags=re.IGNORECASE)
        model = model.strip()
        if model:
            return model
    return None


def parse_loss_datetime(paragraphs: list[str]) -> tuple[date | None, time | None]:
    """
    Parse loss date and time from the 'Дата та час втрати:' line.
    Handles both ':' and '.' as time separators (e.g. 09:50 and 22.36).
    """
    for para in paragraphs:
        if "дата та час втрати" not in para.lower():
            continue

        loss_date: date | None = None
        loss_time: time | None = None

        date_m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", para)
        if date_m:
            try:
                loss_date = date(int(date_m.group(3)), int(date_m.group(2)), int(date_m.group(1)))
            except ValueError:
                pass

            # Search for time only in text that follows the date to avoid false matches
            after_date = para[date_m.end():]
            time_m = re.search(r"(\d{2})[:\.](\d{2})", after_date)
            if time_m:
                try:
                    loss_time = time(int(time_m.group(1)), int(time_m.group(2)))
                except ValueError:
                    pass

        return loss_date, loss_time
    return None, None


def parse_coordinates(paragraphs: list[str]) -> str:
    """
    Extract MGRS/UTM coordinates from the flight-duration paragraph.
    Strips height annotation (h: XXX) to match the output xlsx format.

    Source line example:
      'Зліт о 10:00, 3 хв. польоту, (37U CP 45225 47782 h: 164) Північ...'
    Returns:
      '37U CP 45225 47782'
    """
    for para in paragraphs:
        para_lower = para.lower()
        if "зліт о" not in para_lower and "хв. польоту" not in para_lower:
            continue
        m = re.search(r"\(([^)]+)\)", para)
        if m:
            coords = m.group(1).strip()
            # Remove height annotation: ' h: 154' or ' h:160'
            coords = re.sub(r"\s*h\s*:\s*\d+", "", coords, flags=re.IGNORECASE).strip()
            # Collapse multiple spaces that may remain
            coords = re.sub(r" {2,}", " ", coords)
            return coords
    return ""


def _is_vampire_or_heavy(model: str) -> bool:
    """Return True for Vampire / HeavyShot platforms (case-insensitive)."""
    keywords = ["vampire", "вампір", "вампир", "heavy shot", "heavyshot", "важкий"]
    model_lower = model.lower()
    return any(kw in model_lower for kw in keywords)


def get_frequencies(model: str) -> tuple[float, float]:
    """
    Return (control_freq_GHz, video_freq_GHz) based on drone model name.

    Vampire / HeavyShot  ->  2.4 / 2.4
    All other models     ->  2.4 / 5.8

    NOTE: frequency data is not present in the .docx files. This lookup
    table should be expanded as new drone types appear in reports.
    """
    if _is_vampire_or_heavy(model):
        return 2.4, 2.4
    return 2.4, 5.8


def get_flight_clearance(model: str) -> str:
    """
    Return flight-clearance field value based on drone model.

    Vampire / HeavyShot  ->  'є'  (clearance required and filed)
    All other models     ->  '-'  (прочерк)
    """
    return "є" if _is_vampire_or_heavy(model) else "-"

# ---------------------------------------------------------------------------
# Record assembly
# ---------------------------------------------------------------------------


def parse_drone_loss(paragraphs: list[str], filename: str) -> dict | None:
    """
    Parse all fields from a drone loss report.
    Returns a populated dict on success, or None if mandatory fields are absent.
    """
    model = parse_drone_model(paragraphs)
    if not model:
        log.warning(f"  [УВАГА] Не знайдено модель БпЛА у файлі: {filename}")
        return None

    loss_date, loss_time = parse_loss_datetime(paragraphs)
    if not loss_date:
        log.warning(f"  [УВАГА] Не знайдено дату втрати у файлі: {filename}")
        return None

    unit_raw = _find_inline_or_next(paragraphs, "Військова частина, підрозділ")
    serial_raw = _find_inline_or_next(paragraphs, "Серійний номер")
    reason_raw = _find_inline_or_next(paragraphs, "Короткі відомості")

    freq_control, freq_video = get_frequencies(model)

    return {
        "model": model,
        "loss_date": loss_date,
        "loss_time": loss_time,
        "coordinates": parse_coordinates(paragraphs),
        "freq_control": freq_control,
        "freq_video": freq_video,
        "unit": (unit_raw or "").rstrip(".").strip(),
        "serial": (serial_raw or "").rstrip(".").strip(),
        "loss_reason": (reason_raw or "").strip(),
        # Standard value used in the output xlsx for this column
        "what_happened": "Повна втрата сигналу пульта з дроном",
        "flight_clearance": get_flight_clearance(model),
        "source_file": filename,
    }

# ---------------------------------------------------------------------------
# Output filename
# ---------------------------------------------------------------------------


def generate_output_filename(records: list[dict]) -> str:
    """
    Build the output filename from the latest loss date across all records.
    Pattern: ЗВІТ_втрачені_БпЛА_по_DD_MM_YYYY.xlsx

    NOTE: The example output file also contains a brigade identifier in the name
    (e.g. '3_БрОП_НГУ'). Once reports are standardized to a single brigade,
    extract the brigade name from document content and insert it here.
    """
    dates = [r["loss_date"] for r in records if r.get("loss_date")]
    if dates:
        return f"ЗВІТ_втрачені_БпЛА_по_{max(dates).strftime('%d_%m_%Y')}.xlsx"
    return "ЗВІТ_втрачені_БпЛА.xlsx"

# ---------------------------------------------------------------------------
# XLSX output
# ---------------------------------------------------------------------------

_HEADERS = [
    "№",
    "Назва БпЛА",
    "Час",
    "Дата",
    "Місце втрати (КООРДИНАТИ)",
    "Частота керування, МГц",
    "Частота відеоканалу, МГц",
    "Що відбувалось з БпЛА під час подавлення",
    (
        "НАЯВНІСТЬ ЗАЯВКИ НА ПРОЛЬОТ\n"
        "(перед заповненням поля розібратись)"
    ),
    (
        "Причина втрати\n"
        "(Тех. Проблеми / Погодні умови, Збито (стрілецька зброя/ФПВ противника), "
        " Ворожий РЕБ, Дружній  РЕБ)"
    ),
    "Серійний номер",
    "Військова частина / підрозділ",
]

_HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_HEADER_FONT = Font(bold=True, name="Calibri", size=11)
_CELL_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
# Minimum column widths (characters) — adjust if content is wider
_COL_WIDTHS = [5, 22, 8, 13, 30, 12, 12, 38, 18, 48, 24, 32]


def write_xlsx(records: list[dict], output_path: Path) -> None:
    """Write all parsed records to a formatted .xlsx file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Аркуш1"

    # Header row
    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = _CELL_BORDER
    ws.row_dimensions[1].height = 60

    # Sort by date then time (earliest first)
    sorted_records = sorted(
        records,
        key=lambda r: (r["loss_date"], r["loss_time"] or time(0, 0)),
    )

    # Data rows
    for row_idx, rec in enumerate(sorted_records, start=2):
        row_data = [
            row_idx - 1,           # №
            rec["model"],          # Назва БпЛА
            rec["loss_time"],      # Час  (datetime.time -> Excel time fraction)
            rec["loss_date"],      # Дата (datetime.date -> Excel date serial)
            rec["coordinates"],    # Координати
            rec["freq_control"],   # Частота керування
            rec["freq_video"],     # Частота відео
            rec["what_happened"],  # Що відбувалось
            rec["flight_clearance"],  # Наявність заявки
            rec["loss_reason"],    # Причина втрати
            rec["serial"],         # Серійний номер
            rec["unit"],           # Військова частина
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = _CELL_BORDER
            if col_idx == 3 and value is not None:   # Time
                cell.number_format = "HH:MM"
            elif col_idx == 4 and value is not None: # Date
                cell.number_format = "DD.MM.YYYY"
            elif col_idx in (6, 7):                  # Frequencies
                cell.number_format = "0.0"

    # Column widths
    for col_idx, width in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    wb.save(str(output_path))

# ---------------------------------------------------------------------------
# Summary output
# ---------------------------------------------------------------------------


def print_summary(total: int, success: int, failed: int, output_path: Path | None) -> None:
    log.info("")
    log.info("=" * 60)
    log.info("   РЕЗУЛЬТАТ")
    log.info("=" * 60)
    log.info(f"   Всього файлів:            {total}")
    log.info(f"   Успішно оброблено:        {success}")
    log.info(f"   Не вдалося розпізнати:    {failed}")
    if output_path:
        log.info("")
        log.info(f"   Файл збережено: {output_path.name}")
    log.info("=" * 60)

# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def main() -> None:
    log.info("")
    log.info("=" * 60)
    log.info("   Парсер втрат БпЛА — старт")
    log.info("=" * 60)

    # When frozen by PyInstaller the exe lives next to the .docx files;
    # use its directory rather than cwd (cwd is often ~ on macOS/Windows).
    if getattr(sys, "frozen", False):
        folder = Path(sys.executable).resolve().parent
    else:
        folder = Path.cwd()
    docx_files = find_docx_files(folder)

    if not docx_files:
        log.info("")
        log.info("[УВАГА] Файли .docx не знайдені у поточній папці.")
        log.info(f"        Папка: {folder}")
        log.info("        Переконайтесь, що програма знаходиться разом з файлами донесень.")
        input("\nНатисніть Enter для виходу...")
        return

    log.info(f"[INFO] Знайдено файлів .docx: {len(docx_files)}")

    records: list[dict] = []
    failed = 0

    for filepath in docx_files:
        log.info(f"\n[INFO] Обробка: {filepath.name}")
        try:
            paragraphs = extract_paragraphs(filepath)
            record = parse_drone_loss(paragraphs, filepath.name)
            if record:
                records.append(record)
                time_str = record["loss_time"].strftime("%H:%M") if record["loss_time"] else "??:??"
                date_str = record["loss_date"].strftime("%d.%m.%Y")
                coords_str = record["coordinates"] or "координати не знайдено"
                log.info(f"[OK]   {record['model']} | {date_str} {time_str} | {coords_str}")
            else:
                failed += 1
                log.warning("[ПРОПУСК] Файл пропущено — обов'язкові поля відсутні.")
        except Exception as exc:
            failed += 1
            log.warning(f"[ПОМИЛКА] Не вдалося прочитати файл: {exc}")

    if not records:
        log.info("")
        log.info("[УВАГА] Жодного запису не розпізнано. Файл не створено.")
        print_summary(len(docx_files), 0, failed, None)
        input("\nНатисніть Enter для виходу...")
        return

    output_filename = generate_output_filename(records)
    output_path = folder / output_filename

    try:
        write_xlsx(records, output_path)
    except PermissionError:
        log.error(f"\n[ПОМИЛКА] Не вдалося зберегти файл: {output_filename}")
        log.error("          Можливо, він відкритий у Excel. Закрийте файл і спробуйте ще раз.")
        print_summary(len(docx_files), len(records), failed, None)
        input("\nНатисніть Enter для виходу...")
        return

    print_summary(len(docx_files), len(records), failed, output_path)
    input("\nНатисніть Enter для виходу...")


if __name__ == "__main__":
    main()
